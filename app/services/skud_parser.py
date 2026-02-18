from datetime import date, time, datetime, timedelta
from openpyxl import load_workbook
from ..models import PunchRecord


def parse_skud_xlsx(filepath: str, date_from: date, date_to: date) -> list[PunchRecord]:
    """Parse SKUD XLSX export and return list of PunchRecord objects.

    Reads Employee ID, Date, Time columns.
    Adds 1-day buffer on each side of the date range for night shift pairing.
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Date buffer for night shift pairing across date boundaries
    buffer_from = date_from - timedelta(days=1)
    buffer_to = date_to + timedelta(days=1)

    # Find column indices from header row (row 2)
    header_row = None
    col_map = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=False), start=1):
        cells = [cell.value for cell in row]
        if 'Employee ID' in cells:
            header_row = row_idx
            for i, val in enumerate(cells):
                if val:
                    col_map[str(val).strip()] = i
            break

    if header_row is None:
        wb.close()
        raise ValueError("Could not find header row with 'Employee ID' column")

    emp_col = col_map.get('Employee ID')
    date_col = col_map.get('Date')
    time_col = col_map.get('Time')

    if any(c is None for c in [emp_col, date_col, time_col]):
        wb.close()
        raise ValueError(f"Missing required columns. Found: {list(col_map.keys())}")

    punches = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        emp_id_raw = row[emp_col] if emp_col < len(row) else None
        date_raw = row[date_col] if date_col < len(row) else None
        time_raw = row[time_col] if time_col < len(row) else None

        if emp_id_raw is None or date_raw is None or time_raw is None:
            continue

        emp_id = str(emp_id_raw).strip()

        # Parse date
        if isinstance(date_raw, datetime):
            punch_date = date_raw.date()
        elif isinstance(date_raw, date):
            punch_date = date_raw
        else:
            try:
                punch_date = datetime.strptime(str(date_raw).strip(), '%Y-%m-%d').date()
            except ValueError:
                continue

        # Filter by buffered date range
        if punch_date < buffer_from or punch_date > buffer_to:
            continue

        # Parse time
        if isinstance(time_raw, time):
            punch_time = time_raw
        elif isinstance(time_raw, datetime):
            punch_time = time_raw.time()
        else:
            try:
                punch_time = datetime.strptime(str(time_raw).strip(), '%H:%M:%S').time()
            except ValueError:
                continue

        punch_datetime = datetime.combine(punch_date, punch_time)
        punches.append(PunchRecord(
            employee_id=emp_id,
            punch_date=punch_date,
            punch_time=punch_time,
            punch_datetime=punch_datetime,
        ))

    wb.close()
    return punches
